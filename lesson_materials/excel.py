from openpyxl import Workbook, load_workbook

# === Часть 1: Создание нового Excel-файла ===
def create_excel_file(filename):
    # Создаем новый рабочий файл
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Пример"

    # Записываем данные в ячейки
    sheet["A1"] = "Имя"
    sheet["B1"] = "Возраст"
    sheet["C1"] = "Город"

    data = [
        ("Иван", 25, "Москва"),
        ("Анна", 30, "Санкт-Петербург"),
        ("Петр", 35, "Екатеринбург"),
    ]

    for row in data:
        sheet.append(row)

    # Сохраняем файл
    workbook.save(filename)
    print(f"Файл {filename} создан.")


# === Часть 2: Чтение данных из Excel-файла ===
def read_excel_file(filename):
    workbook = load_workbook(filename)
    sheet = workbook.active

    # Читаем данные из таблицы
    print(f"Чтение данных из файла {filename}:")
    for row in sheet.iter_rows(values_only=True):
        print(row)


# === Часть 3: Изменение существующего Excel-файла ===
def modify_excel_file(filename):
    workbook = load_workbook(filename)
    sheet = workbook.active

    # Изменяем данные
    sheet["B2"] = 26  # Меняем возраст Ивана
    sheet["C3"] = "Новосибирск"  # Меняем город Петра

    # Добавляем новый столбец
    sheet["D1"] = "Статус"
    sheet["D2"] = "Активный"
    sheet["D3"] = "Неактивный"
    sheet["D4"] = "Активный"

    # Сохраняем изменения
    workbook.save(filename)
    print(f"Файл {filename} изменен.")


# === Часть 4: Основной код ===
if __name__ == "__main__":
    file_name = "example.xlsx"

    # Создание нового файла
    create_excel_file(file_name)

    # Чтение данных из файла
    read_excel_file(file_name)

    # Изменение данных в файле
    modify_excel_file(file_name)

    # Проверка изменений
    read_excel_file(file_name)
