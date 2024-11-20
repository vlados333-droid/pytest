from openpyxl import Workbook, load_workbook


def create_excel_file(filename):
    work_book = Workbook()
    sheet = work_book.active
    sheet.title = 'Example'

    sheet['A1'] = 'Name'
    sheet['B1'] = 'Age'
    sheet['C1'] = 'City'

    data = [
        ('Иван', 25, "Москва"),
        ('Анна', 30, "Астана"),
        ('Петр', 35, "Алматы"),
    ]

    for row in data:
        sheet.append(row)

    work_book.save(filename)
    print(f"Файл {filename} создан")


def read_excel_file(filename):
    work_book = load_workbook(filename)
    sheet = work_book.active

    print(f'Чтение данных из файла {filename}: ')
    for row in sheet.iter_rows(values_only=True):
        print(row)


def modify_excel_file(filename):
    work_book = load_workbook(filename)
    sheet = work_book.active

    sheet['B2'] = 26

    sheet['D1'] = 'Status'
    sheet['D2'] = 'active'
    sheet['D3'] = 'active'
    sheet['D4'] = 'no active'

    work_book.save(filename)
    print(f'Файл {filename} именён')


if __name__ == "__main__":
    file_name = 'example.xlsx'

    create_excel_file(file_name)
    read_excel_file(file_name)
    modify_excel_file(file_name)
    read_excel_file(file_name)
